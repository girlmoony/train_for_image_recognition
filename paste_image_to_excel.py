import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import os

# 設定
excel_path = "対象のエクセルファイル.xlsx"
画像フォルダ = "画像のあるフォルダパス"

# pandasでエクセル読み込み
df = pd.read_excel(excel_path)

# openpyxlでエクセルを開く
wb = load_workbook(excel_path)
ws = wb.active

# 画像名列のインデックス取得（pandasは0始まり、openpyxlは1始まり）
画像名列_index = df.columns.get_loc("画像名") + 1
画像貼り付け列_index = 画像名列_index + 1

# 画像名列の次に列を挿入
ws.insert_cols(画像貼り付け列_index)

# 画像列のヘッダー名を設定（1行目）
ws.cell(row=1, column=画像貼り付け列_index, value="画像")

# 各行に画像を貼り付け
for i, row in df.iterrows():
   判定結果 = row["判定結果"]

    if 判定結果 == "〇/×": 
    img_name = row["画像名"]
    img_path = os.path.join(画像フォルダ, img_name)



    if os.path.exists(img_path):
        img = ExcelImage(img_path)
        
        # 必要ならサイズ調整
        img.width = 100
        img.height = 100

        # セル位置計算（データ行は2行目から）
        cell_pos = ws.cell(row=i+2, column=画像貼り付け列_index).coordinate
        img.anchor = cell_pos

        ws.add_image(img)

# 保存
wb.save(excel_path)
print("画像列を挿入して画像貼り付け完了")

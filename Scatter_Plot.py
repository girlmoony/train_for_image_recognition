pip install plotly kaleido openpyxl

import plotly.express as px
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

df = pd.read_excel('classification_scores.xlsx')

def plotly_and_save(x_col, y_col, label_col, title, img_filename):
    fig = px.scatter(df, x=x_col, y=y_col, hover_name=label_col, title=title)
    fig.update_traces(marker=dict(size=5), text=None)
    fig.write_image(img_filename)  # 保存（静的画像）

# 画像保存
plotly_and_save('ipro_f1', 'gpu_f1', 'class', 'F1 Score Comparison', 'f1.png')
plotly_and_save('ipro_precision', 'gpu_precision', 'class', 'Precision Comparison', 'precision.png')
plotly_and_save('ipro_recall', 'gpu_recall', 'class', 'Recall Comparison', 'recall.png')

# Excelへの貼り付け（同じ）
wb = load_workbook('classification_scores.xlsx')
ws = wb.create_sheet('Graphs') if 'Graphs' not in wb.sheetnames else wb['Graphs']
img_files = ['f1.png', 'precision.png', 'recall.png']
positions = ['A1', 'A20', 'A40']
for img_file, pos in zip(img_files, positions):
    img = ExcelImage(img_file)
    ws.add_image(img, pos)
wb.save('classification_scores.xlsx')


import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

# Excelファイル読み込み
df = pd.read_excel('classification_scores.xlsx')

# 画像保存とExcel貼り付け用の関数
def plot_and_save(x_col, y_col, label_col, title, img_filename):
    # 散布図作成
    plt.figure(figsize=(6, 6))
    plt.scatter(df[x_col], df[y_col])
    plt.plot([0, 1], [0, 1], 'r--')
    for i in range(len(df)):
        plt.text(df[x_col][i] + 0.005, df[y_col][i], str(df[label_col][i]))
    plt.xlabel(x_col)
    plt.ylabel(y_col)
    plt.title(title)
    plt.grid(True)
    plt.tight_layout()
    
    # 画像として保存
    plt.savefig(img_filename)
    plt.close()

# 3つの図を保存
plot_and_save('ipro_f1', 'gpu_f1', 'class', 'F1 Score Comparison', 'f1.png')
plot_and_save('ipro_precision', 'gpu_precision', 'class', 'Precision Comparison', 'precision.png')
plot_and_save('ipro_recall', 'gpu_recall', 'class', 'Recall Comparison', 'recall.png')

# Excelファイルに画像を貼り付ける
wb = load_workbook('classification_scores.xlsx')
if 'Graphs' not in wb.sheetnames:
    ws = wb.create_sheet('Graphs')
else:
    ws = wb['Graphs']

# 各画像をExcelシートに貼り付け
img_files = ['f1.png', 'precision.png', 'recall.png']
positions = ['A1', 'A20', 'A40']  # 位置調整

for img_file, pos in zip(img_files, positions):
    img = ExcelImage(img_file)
    ws.add_image(img, pos)

# 保存
wb.save('classification_scores.xlsx')



import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

# Excelファイル読み込み
excel_path = 'classification_scores.xlsx'
df = pd.read_excel(excel_path)

# 指標リスト
metrics = ['f1', 'precision', 'recall']
img_files = []

# 散布図を作成・保存
for metric in metrics:
    ipro_col = f'ipro_{metric}'
    gpu_col = f'gpu_{metric}'
    
    # 相関係数計算
    corr = np.corrcoef(df[ipro_col], df[gpu_col])[0, 1]
    
    # グラフ描画
    plt.figure(figsize=(6, 6))
    plt.scatter(df[ipro_col], df[gpu_col])
    plt.plot([0, 1], [0, 1], 'r--', label='y = x')
    for i in range(len(df)):
        plt.text(df[ipro_col][i] + 0.005, df[gpu_col][i], str(df['class'][i]), fontsize=8)
    plt.xlabel(f'iPro {metric.capitalize()}')
    plt.ylabel(f'GPU {metric.capitalize()}')
    plt.title(f'{metric.capitalize()} Comparison\n(Pearson r = {corr:.3f})')
    plt.grid(True)
    plt.legend()
    plt.tight_layout()
    
    # 保存
    filename = f'{metric}_comparison.png'
    plt.savefig(filename)
    plt.close()
    img_files.append(filename)

# Excelファイルに画像を貼り付け
wb = load_workbook(excel_path)
ws = wb.create_sheet('Graphs') if 'Graphs' not in wb.sheetnames else wb['Graphs']

# 各画像を適切な位置に貼る
positions = ['A1', 'A20', 'A40']  # 位置調整（縦に並べる）
for img_file, pos in zip(img_files, positions):
    img = ExcelImage(img_file)
    ws.add_image(img, pos)

# 保存
wb.save(excel_path)


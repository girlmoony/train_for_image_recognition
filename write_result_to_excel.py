import cv2
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os
from sklearn.preprocessing import LabelEncoder
from skimage import feature
import seaborn as sns

# 図の保存先フォルダ
output_img_dir = "output_images"
os.makedirs(output_img_dir, exist_ok=True)

# Excel出力ファイル
output_excel = "output_per_class.xlsx"

# ラベルファイル読み込み
labels_df = pd.read_excel("your_labels.xlsx", sheet_name="image_details")

# 結果をリストに格納
data = []

# 特徴量抽出関数
def extract_features(image_path):
    img = cv2.imread(image_path)
    if img is None:
        print(f"画像読み込み失敗: {image_path}")
        return None

    img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    img_hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    saturation = np.mean(img_hsv[:, :, 1])
    noise_level = cv2.Laplacian(gray, cv2.CV_64F).var()
    gray_8bit = cv2.resize(gray, (64, 64))
    glcm = feature.greycomatrix(gray_8bit, distances=[1], angles=[0], symmetric=True, normed=True)
    texture_energy = feature.greycoprops(glcm, 'energy')[0, 0]
    r_mean = np.mean(img_rgb[:, :, 0])
    g_mean = np.mean(img_rgb[:, :, 1])
    b_mean = np.mean(img_rgb[:, :, 2])
    color_temp = b_mean / (r_mean + 1e-5)

    return saturation, noise_level, texture_energy, color_temp

# データ収集ループ
for idx, row in labels_df.iterrows():
    img_path = row["image_path"]
    true_label = row["true_label"]

    if not os.path.exists(img_path):
        print(f"ファイルが存在しません: {img_path}")
        continue

    features = extract_features(img_path)
    if features is None:
        continue

    saturation, noise_level, texture_energy, color_temp = features

    data.append({
        "image_path": img_path,
        "true_label": true_label,
        "saturation": saturation,
        "noise_level": noise_level,
        "texture_energy": texture_energy,
        "color_temp": color_temp
    })

# DataFrame化
df = pd.DataFrame(data)
print(df.head())

# Excel書き込み
with pd.ExcelWriter(output_excel) as writer:
    for class_name in df["true_label"].unique():
        sub_df = df[df["true_label"] == class_name]
        sub_df.to_excel(writer, sheet_name=str(class_name), index=False)

        # 図保存
        feature_list = ["saturation", "noise_level", "texture_energy", "color_temp"]
        for feature in feature_list:
            plt.figure(figsize=(8, 5))
            sns.boxplot(data=sub_df, x="true_label", y=feature)
            plt.title(f"{class_name} - {feature}")
            plt.tight_layout()

            img_save_path = os.path.join(output_img_dir, f"{class_name}_{feature}.png")
            plt.savefig(img_save_path)
            plt.close()

print(f"\nExcel保存先: {output_excel}")
print(f"図の保存先フォルダ: {output_img_dir}")



import openpyxl
from openpyxl.drawing.image import Image as XLImage
import cv2
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os
from skimage import feature
import seaborn as sns

# 出力ファイル・フォルダ
output_excel = "output_with_images.xlsx"
output_img_dir = "output_images"
os.makedirs(output_img_dir, exist_ok=True)

# ラベルファイル
labels_df = pd.read_excel("your_labels.xlsx", sheet_name="image_details")

# 特徴量抽出関数（省略。必要なら前述コードを入れてください）

# 特徴量収集（省略。必要なら前述コードを入れてください）

# データFrame化
df = pd.DataFrame(data)

# Excelブック作成
wb = openpyxl.Workbook()
del wb["Sheet"]  # デフォルトシート削除

# 特徴量リスト
feature_list = ["saturation", "noise_level", "texture_energy", "color_temp"]

# 各クラスごとに処理
for class_name in df["true_label"].unique():
    sub_df = df[df["true_label"] == class_name]

    # 図生成＆保存
    img_paths = []
    for feature in feature_list:
        plt.figure(figsize=(8, 5))
        sns.boxplot(data=sub_df, x="true_label", y=feature)
        plt.title(f"{class_name} - {feature}")
        plt.tight_layout()

        img_path = os.path.join(output_img_dir, f"{class_name}_{feature}.png")
        plt.savefig(img_path)
        plt.close()

        img_paths.append(img_path)

    # シート作成
    ws = wb.create_sheet(title=str(class_name))

    # 画像貼り付け
    row_offset = 1
    for img_path in img_paths:
        img = XLImage(img_path)
        ws.add_image(img, f"A{row_offset}")

        # 適当に行間隔を空ける（画像サイズにより調整必要）
        row_offset += 20

# 保存
wb.save(output_excel)
print(f"画像付きExcelを保存しました: {output_excel}")


import openpyxl
from openpyxl.drawing.image import Image as XLImage
import cv2
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os
from sklearn.preprocessing import LabelEncoder
from sklearn.ensemble import RandomForestClassifier
from skimage import feature
import seaborn as sns

# 出力先
output_excel = "output_per_class_feature_importance.xlsx"
output_img_dir = "output_images"
os.makedirs(output_img_dir, exist_ok=True)

# ラベルファイル読み込み
labels_df = pd.read_excel("your_labels.xlsx", sheet_name="image_details")

# 特徴量抽出関数（省略、必要なら前述のものを使ってください）

# 特徴量収集
data = []
for idx, row in labels_df.iterrows():
    img_path = row["image_path"]
    true_label = row["true_label"]

    if not os.path.exists(img_path):
        continue

    features = extract_features(img_path)
    if features is None:
        continue

    saturation, noise_level, texture_energy, color_temp = features

    data.append({
        "image_path": img_path,
        "true_label": true_label,
        "saturation": saturation,
        "noise_level": noise_level,
        "texture_energy": texture_energy,
        "color_temp": color_temp
    })

df = pd.DataFrame(data)

# Excelブック準備
wb = openpyxl.Workbook()
del wb["Sheet"]

# 特徴量リスト
feature_list = ["saturation", "noise_level", "texture_energy", "color_temp"]

# クラスごとに処理
for class_name in df["true_label"].unique():
    sub_df = df[df["true_label"] == class_name]

    if sub_df.shape[0] < 5:
        print(f"{class_name}のデータが少なすぎるためスキップ")
        continue

    # ラベルエンコード（このクラス内のラベルで）
    le = LabelEncoder()
    sub_df["label_encoded"] = le.fit_transform(sub_df["true_label"])

    # 特徴量とラベル
    X_sub = sub_df[feature_list].values
    y_sub = sub_df["label_encoded"].values

    # モデル学習
    model = RandomForestClassifier(n_estimators=100, random_state=42)
    model.fit(X_sub, y_sub)

    # 重要度取得
    importances = model.feature_importances_
    importance_df = pd.DataFrame({
        "Feature": feature_list,
        "Importance": importances
    }).sort_values(by="Importance", ascending=False)

    # 図生成・保存
    img_paths = []
    for feature in feature_list:
        plt.figure(figsize=(8, 5))
        sns.boxplot(data=sub_df, x="true_label", y=feature)
        plt.title(f"{class_name} - {feature}")
        plt.tight_layout()

        img_path = os.path.join(output_img_dir, f"{class_name}_{feature}.png")
        plt.savefig(img_path)
        plt.close()

        img_paths.append(img_path)

    # シート作成
    ws = wb.create_sheet(title=str(class_name))

    # 重要度書き込み
    ws.cell(row=1, column=1, value="重要特徴量（このクラス内）")
    for i, (feat, imp) in enumerate(zip(importance_df["Feature"], importance_df["Importance"])):
        ws.cell(row=i + 2, column=1, value=feat)
        ws.cell(row=i + 2, column=2, value=float(imp))

    # 画像貼り付け
    row_offset = len(importance_df) + 4
    for img_path in img_paths:
        img = XLImage(img_path)
        ws.add_image(img, f"A{row_offset}")
        row_offset += 20

wb.save(output_excel)
print(f"クラスごとに個別の重要特徴量と画像を含めたExcelを保存しました： {output_excel}")


